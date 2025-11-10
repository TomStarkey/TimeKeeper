import os

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


# Takes Date, and Time worked from gui.
# Loads workbook and appends this data to 2 columns in Excel
# If file does not already exist it will create it.


class Gatherer:
    def __init__(self):
        self.time_worked = ""
        self.date = ""
        self.time = ""
        self.label = ""

    def recieve(self, date: str, time: str, label: str):
        self.time_worked = time
        self.date = date
        self.label = label

    def write(self):
        excel_file = "TimeWorked.xlsx"
        if not os.path.isfile(excel_file):
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Date", "Time", "Time worked", "Label"])
            sheet.append([self.date, self.time, self.time_worked, self.label])
        else:
            wb = load_workbook(excel_file)
            sheet = wb.active
            sheet.append([self.date, self.time, self.time_worked, self.label])

        wb.save(excel_file)

    def reset(self):
        self.time_worked = ""
        self.date = ""
        self.label = ""

    def split_datetime(self):
        date = str(self.date)
        comma_idx = date.index(",")

        time = date[:comma_idx]
        date = date[comma_idx+1:]

        self.date = date
        self.time = time